import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from export_utils import export_room_usage


def test_export_room_usage_handles_blank_and_filtered_room_names():
    df = pd.DataFrame([
        {"grade": 1, "class": 1, "day": "月", "period": 1, "subject": "Math", "teacher": "T1", "room": ""},
        {"grade": 1, "class": 1, "day": "火", "period": 1, "subject": "Sci", "teacher": "T2", "room": "???"},
    ])
    data = export_room_usage(df)
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["raw_data", "room1", "room2"]
    assert len(set(wb.sheetnames)) == len(wb.sheetnames)
