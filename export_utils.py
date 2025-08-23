"""Utility functions for exporting timetable data to Excel files.

Each export helper returns ``bytes`` representing an Excel workbook. In
addition to the requested summary sheets, every workbook now contains a
``raw_data`` sheet with the original timetable ``DataFrame`` so that users
can inspect or further process the underlying data easily.

The module provides exports for class schedules, teacher schedules,
room usage and subject summaries.
"""

import pandas as pd
from io import BytesIO
import re
from openpyxl.styles import Border, Side
import zipfile

# Common ordering for days of the week and periods so that all exports share
# the same layout. Previously these were imported from the streamlit app which
# caused a ``NameError`` when the functions were used independently.  Defining
# them here keeps the utilities self‑contained.
WEEK = ["月", "火", "水", "木", "金"]
PERIODS = [1, 2, 3, 4, 5, 6]

def export_class_schedule(df: pd.DataFrame) -> bytes:
    """Create an Excel file of class schedules from a timetable DataFrame.

    Parameters
    ----------
    df: DataFrame
        Timetable with columns [grade, class, day, period, subject, teacher, room].

    Returns
    -------
    bytes
        Excel file bytes where each sheet corresponds to a class and contains a
        pivot table with days as columns and periods as rows.  For each period
        three consecutive rows are used – subject, teacher and room – so that
        timetable details occupy separate rows instead of being combined into a
        single cell.
    """
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Include raw data sheet for reference
        df.to_excel(writer, sheet_name="raw_data", index=False)

        for (grade, cls), group in df.groupby(["grade", "class"]):
            pivots = {}
            for col in ["subject", "teacher", "room"]:
                piv = (
                    group.pivot_table(
                        index="period",
                        columns="day",
                        values=col,
                        aggfunc=lambda x: "\n".join(x),
                        fill_value="",
                    )
                    .reindex(index=PERIODS, columns=WEEK, fill_value="")
                )
                pivots[col] = piv

            combined = pd.concat(pivots.values(), keys=pivots.keys())
            combined = combined.swaplevel(0, 1)
            combined.index.names = ["period", "info"]
            combined = combined.reindex(
                pd.MultiIndex.from_product(
                    [PERIODS, ["subject", "teacher", "room"]],
                    names=["period", "info"],
                ),
                fill_value="",
            )

            display_df = combined.reset_index()
            # period 表示は 3 行 1 セットの先頭（=subject の行）のみ
            display_df.loc[display_df["info"] != "subject", "period"] = ""
            display_df = display_df.drop(columns=["info"])

            sheet_name = f"{grade}-{cls}"
            display_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # ===== ここから枠線の付与 =====
            ws = writer.sheets[sheet_name]
            thin = Side(style="thin", color="000000")

            n_body_rows = display_df.shape[0]       # データ行数（ヘッダ除く）
            n_cols = display_df.shape[1]            # 列数（period + 曜日列）
            n_total_rows = n_body_rows + 1          # ヘッダ込み

            # すべてのセルに「列の右側」の枠線 → 曜日ごとの区切り線になる
            for r in range(1, n_total_rows + 1):        # 1 はヘッダ行
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border
                    cell.border = Border(
                        left=b.left,
                        right=thin,   # 列の右側に縦線
                        top=b.top,
                        bottom=b.bottom,
                    )

            # ヘッダ行の下に横線
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=c)
                b = cell.border
                cell.border = Border(
                    left=b.left,
                    right=b.right,
                    top=b.top,
                    bottom=thin,
                )

            # period ブロック（3行1セット）の区切りに横線（下枠線）
            # データは 2 行目から始まるので、i=0 の行は Excel の row=2
            for i in range(n_body_rows):
                if (i + 1) % 3 == 0:  # 各ブロックの末尾行
                    r = 2 + i
                    for c in range(1, n_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        b = cell.border
                        cell.border = Border(
                            left=b.left,
                            right=b.right,   # 既存の縦線は維持
                            top=b.top,
                            bottom=thin,     # ブロックの下に横線
                        )
    buffer.seek(0)
    return buffer.getvalue()

def export_teacher_schedule(df: pd.DataFrame) -> bytes:
    """Create an Excel file of teacher schedules.

    For each teacher a sheet is created with days as rows and periods as
    columns.  Within each period three consecutive rows are used to show
    the subject, teacher and room so that details are separated instead of
    combined into a single cell.

    Parameters
    ----------
    df : pandas.DataFrame
        Scheduling table containing at least the columns
        ``['grade', 'class', 'day', 'period', 'subject', 'teacher', 'room']``.

    Returns
    -------
    bytes
        Excel file data.
    """

    expanded = df.copy()
    expanded = expanded[expanded['teacher'] != '']
    expanded = expanded.assign(teacher=expanded['teacher'].str.split('/')).explode('teacher')

    # combine class and subject for display in the subject row
    expanded['subject_line'] = (
        expanded['grade'].astype(str)
        + '-' + expanded['class'].astype(str)
        + ' ' + expanded['subject']
    )

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="raw_data", index=False)

        for teacher, group in expanded.groupby('teacher'):
            pivots = {}
            for col, key in [("subject_line", "subject"), ("teacher", "teacher"), ("room", "room")]:
                piv = (
                    group.pivot_table(
                        index="day",
                        columns="period",
                        values=col,
                        aggfunc=lambda x: "\n".join(x),
                        fill_value="",
                    )
                    .reindex(index=WEEK, columns=PERIODS, fill_value="")
                )
                pivots[key] = piv

            combined = pd.concat(pivots.values(), keys=pivots.keys())
            combined = combined.swaplevel(0, 1)
            combined.index.names = ["day", "info"]
            combined = combined.reindex(
                pd.MultiIndex.from_product(
                    [WEEK, ["subject", "teacher", "room"]],
                    names=["day", "info"],
                ),
                fill_value="",
            )

            display_df = combined.reset_index()
            display_df.loc[display_df["info"] != "subject", "day"] = ""
            display_df = display_df.drop(columns=["info"])

            safe_name = re.sub(r'[\\/*?\[\]:]', '', teacher)[:31]
            display_df.to_excel(writer, sheet_name=safe_name, index=False)

            ws = writer.sheets[safe_name]
            thin = Side(style="thin", color="000000")

            n_body_rows = display_df.shape[0]
            n_cols = display_df.shape[1]
            n_total_rows = n_body_rows + 1

            for r in range(1, n_total_rows + 1):
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border
                    cell.border = Border(
                        left=b.left,
                        right=thin,
                        top=b.top,
                        bottom=b.bottom,
                    )

            for c in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=c)
                b = cell.border
                cell.border = Border(
                    left=b.left,
                    right=b.right,
                    top=b.top,
                    bottom=thin,
                )

            for i in range(n_body_rows):
                if (i + 1) % 3 == 0:
                    r = 2 + i
                    for c in range(1, n_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        b = cell.border
                        cell.border = Border(
                            left=b.left,
                            right=b.right,
                            top=b.top,
                            bottom=thin,
                        )

    buffer.seek(0)
    return buffer.getvalue()

def export_room_usage(df: pd.DataFrame) -> bytes:
    """Export room usage as an Excel file with one sheet per room.

    Each sheet lays out days in rows and periods in columns with three
    consecutive rows per day showing subject, teacher and room.  This
    mirrors the layout of :func:`export_class_schedule`.
    """

    df = df.copy()
    df["subject_line"] = (
        df["grade"].astype(str)
        + '-' + df["class"].astype(str)
        + ' ' + df["subject"]
    )

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="raw_data", index=False)

        used_names = set()
        fallback_counter = 1

        for room, group in df.groupby("room"):
            pivots = {}
            for col, key in [("subject_line", "subject"), ("teacher", "teacher"), ("room", "room")]:
                piv = (
                    group.pivot_table(
                        index="day",
                        columns="period",
                        values=col,
                        aggfunc=lambda x: "\n".join(x),
                        fill_value="",
                    )
                    .reindex(index=WEEK, columns=PERIODS, fill_value="")
                )
                pivots[key] = piv

            combined = pd.concat(pivots.values(), keys=pivots.keys())
            combined = combined.swaplevel(0, 1)
            combined.index.names = ["day", "info"]
            combined = combined.reindex(
                pd.MultiIndex.from_product(
                    [WEEK, ["subject", "teacher", "room"]],
                    names=["day", "info"],
                ),
                fill_value="",
            )

            display_df = combined.reset_index()
            display_df.loc[display_df["info"] != "subject", "day"] = ""
            display_df = display_df.drop(columns=["info"])

            safe_name = re.sub(r'[\\/*?\[\]:]', '', str(room)).strip()[:31]
            if not safe_name:
                while True:
                    safe_name = f"room{fallback_counter}"
                    fallback_counter += 1
                    if safe_name not in used_names:
                        break

            base_name = safe_name
            suffix = 1
            while safe_name in used_names:
                safe_name = f"{base_name}_{suffix}"
                suffix += 1

            used_names.add(safe_name)

            display_df.to_excel(writer, sheet_name=safe_name, index=False)

            ws = writer.sheets[safe_name]
            thin = Side(style="thin", color="000000")

            n_body_rows = display_df.shape[0]
            n_cols = display_df.shape[1]
            n_total_rows = n_body_rows + 1

            for r in range(1, n_total_rows + 1):
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    b = cell.border
                    cell.border = Border(
                        left=b.left,
                        right=thin,
                        top=b.top,
                        bottom=b.bottom,
                    )

            for c in range(1, n_cols + 1):
                cell = ws.cell(row=1, column=c)
                b = cell.border
                cell.border = Border(
                    left=b.left,
                    right=b.right,
                    top=b.top,
                    bottom=thin,
                )

            for i in range(n_body_rows):
                if (i + 1) % 3 == 0:
                    r = 2 + i
                    for c in range(1, n_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        b = cell.border
                        cell.border = Border(
                            left=b.left,
                            right=b.right,
                            top=b.top,
                            bottom=thin,
                        )

    buffer.seek(0)
    return buffer.getvalue()

def export_subject_summary(df: pd.DataFrame) -> bytes:
    """Aggregate timetable by grade, class and subject and export as Excel.

    The resulting workbook contains two sheets:
    ``summary`` – the aggregated subject counts, and ``raw_data`` – the
    original timetable.
    """

    summary = (
        df.groupby(["grade", "class", "subject"])
        .size()
        .reset_index(name="periods_per_week")
        .sort_values(["grade", "class", "subject"])
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        df.to_excel(writer, sheet_name="raw_data", index=False)

    output.seek(0)
    return output.getvalue()


def export_all_excel(df: pd.DataFrame) -> bytes:
    """Package key Excel exports into a single ZIP archive.

    The archive contains class schedules, teacher schedules and room usage
    workbooks. Each workbook already includes a ``raw_data`` sheet.
    """

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w") as zf:
        zf.writestr("class_schedule.xlsx", export_class_schedule(df))
        zf.writestr("teacher_schedule.xlsx", export_teacher_schedule(df))
        zf.writestr("room_usage.xlsx", export_room_usage(df))
    buffer.seek(0)
    return buffer.getvalue()
